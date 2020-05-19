# -*- coding: utf-8 -*-
import os
import sys
import PyQt5
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5 import uic # ui designer
from modules.file_sort2 import *
import openpyxl

EDLUI = '../_uiFiles/EDLui.ui' # 경로설정

class MainDialog(QDialog):


    def __init__(self):
        QDialog.__init__(self , None)
        uic.loadUi(EDLUI , self)
        self.pushButton_File_1.clicked.connect(self.get_file) #file part
        self.pushButton_File_2.clicked.connect(self.get_file2)
        self.pushButton_run.clicked.connect(self.runButton)
#        self.pushButton.clicked.connect(self.get_cell_name)

    def get_file(self):
        fileURL_1 = QFileDialog.getOpenFileName() # 파일 창이 나오는 함수
        self.textEdit_File_1.setText(fileURL_1[0])
        filename_1 = os.path.basename(fileURL_1[0])
        get_file_1(filename_1 , fileURL_1[0])
        wb = openpyxl.load_workbook(fileURL_1[0] , read_only = True)
        self.get_cell(wb.get_sheet_names() , 0)

    def get_file2(self):
        fileURL_2 = QFileDialog.getOpenFileName()
        self.textEdit_File_2.setText(fileURL_2[0])
        filename_2 = os.path.basename(fileURL_2[0])
        get_file_2(filename_2 , fileURL_2[0])
        wb = openpyxl.load_workbook(fileURL_2[0], read_only=True)
        self.get_cell(wb.get_sheet_names() , 1)

    def get_cell(self , sheetnames, g):
        if g == 0:
            self.comboBox_File_1.addItems(sheetnames)

        else:
            self.comboBox_File_2.addItems(sheetnames)

    def runButton(self):
        saveurl = QFileDialog.getExistingDirectory()
        get_sheetname_1(self.comboBox_File_1.currentText())
        get_sheetname_2(self.comboBox_File_2.currentText())
        get_saveurl(saveurl)
        get_N(self.spinBox.value)
     #  exrun()

        runaway() # 실행

    def get_cell_name(self):
        print(self.comboBox_File_1.currentText())
        print(self.comboBox_File_2.currentText())


if __name__ == "__main__" :
    app = QApplication(sys.argv) # 프로그램 실행시킴
    main_dialog = MainDialog() # 객체 생성
    main_dialog.show() # 화면을 띄우기
    app.exec() # 프그램을 이벤트루프로 진행시킴

# 오류 ! : 실재하지 않는 폴더 클릭 시 에러메세지 출력됨
#경로 중 파일명만 얻기
#os.path.basename(path)