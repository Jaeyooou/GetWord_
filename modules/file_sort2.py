import pandas as pd
import os
import re
from openpyxl import Workbook
from modules.data_label import *
from modules.uimain import *

#문서를 디렉터리

base_dir = ".."
excel_file1 = ""
excel_file2 = ""
excel_dir1 = ""
excel_dir2 = ""
sheet1 = ""
sheet2 = ""
saveURL = ""
N = int()

def get_file_1(filename , fileurl):
    global excel_file1
    global excel_dir1
    excel_file1 = filename
    excel_dir1 = fileurl
def get_file_2(filename , fileurl):
    global excel_file2
    global excel_dir2
    excel_file2 = filename
    excel_dir2 = fileurl
def get_sheetname_1(sheet):
    global sheet1
    sheet1 = sheet
def get_sheetname_2(sheet):
    global sheet2
    sheet2 = sheet
def get_saveurl(saveurl):
    global saveURL
    saveURL = saveurl
def get_N(n):
    global N
    N = n

# def exrun():
#     print(base_dir)
#     print(excel_file1)
#     print(excel_file2)
#     print(excel_dir1)
#     print(excel_dir2)
#     print(sheet1)
#     print(sheet2)


def runaway():
    print("runaway start!")
    df_from_excel1 = pd.read_excel(excel_dir1, sheet_name= sheet1)
    df_from_excel2 = pd.read_excel(excel_dir2, sheet_name= sheet2)
    File_Name = df_from_excel1['"문서명"']  # row_data 는  문서명
    Position_Data = df_from_excel1['"최종결재자직위구분"']
    Step_1_Data = df_from_excel2['1차유형']  # row_data 는  문서명
    Step_2_Data = df_from_excel2['2차유형']
    Rank_Data = df_from_excel2['등급']

    # 데이터 쓰기
    wb = Workbook()  # Workbook()메소드를 이용한 데이터 기입
    ws = wb.active
    ws['A1'] = '문서명'  # header
    ws['B1'] = '1차 최종'

    for i in range(N):
        ws.cell(row=i + 2, column=1).value = File_Name[i]  # 문서명 기입
        '''for w in Position:
            if w in str(Position_Data[i]):
                ws.cell(row=i + 2, column=3).value = w'''
        Only_Char = (re.sub(r'\([^)]*\)', '', File_Name[i]))
        a = []

        for w in Com_Label:  # 1차유형에 있는 모든 단어를 추출
            if w in Only_Char:  # 1차유형에 있는 모든 단어를 추출
                a.append(w)  # 1차유형에 있는 모든 단어를 추출

        if len(a) == 1 and str(a) in Com_Label1:  # 추출한 단어가 1개이고, 1차 유형에서 이미 등급이 정해졌다면
            ws.cell(row=i + 2, column=2).value = a[0]  # 그 단어와 등급을 기입
            ws.cell(row=i + 2, column=4).value = Com_Label1[a[0]]
            continue
        else:  # 추출한 단어가 1개 이상 혹은 등급이 정해지지 않았다면
            dic = dict()  # dic은 1차유형 단어의 딕셔너리
            dic1 = dict()  # dic1은 2차유형 단어의 딕셔너리
            for k in a:
                dic[k] = Only_Char.find(k)  # 뒷쪽에 있는 단어를 추출하기 위해 find 메소드를 활용
            dic_reverse = {v: k for k, v in dic.items()}  # 딕셔너리의 형태를 바꿔줌
            value = ""  # 맨 뒷쪽의 단어를 저장하기 위한 문자열
            b = []
            for k, v in sorted(dic_reverse.items()):
                value = v
                ws.cell(row=i + 2, column=2).value = v  # 맨 마지막 단어를 엑셀에 저장

            if value in Com_Label1:  # 맨 마지막 단어가 등급이 정해져있다면
                ws.cell(row=i + 2, column=4).value = Com_Label1[value]  # 등급 또 엑셀에 저장
                continue

            else:  # 맨 마지막 단어의 등급이 Com_Label1 즉 1차유형에서 등급이 정해져 있는 단어가
                for j in range(len(Step_1_Data)):  # 아니라면 2차유형을 통해 등급을 나누도록 함.
                    if str(value) == Step_1_Data[j] and str(Step_2_Data[j]) in Only_Char:  # 문서명의 맨마지막 단어가 1차유형에 있고
                        b.append(Step_2_Data[j])  # 2차 유형의 단어 또한 문서명에 있다면 그 단어를 리스트에 저장
                if len(b) == 1:  # 2차유형에서 단 한개만이 중복된다면
                    ws.cell(row=i + 2, column=2).value = str(b[0]) + value  # 그 단어를 엑셀에 저장
                    for j in range(len(Step_2_Data)):  # 등급 또한 저장
                        if Step_1_Data[j] == value and Step_2_Data[j] == b[0]:
                            if len(str(Rank_Data[j])) != 1:
                                print(Only_Char, Rank_Data[j])
                                ws.cell(row=i + 2, column=4).value = sep_specific(str(Only_Char), str(Step_1_Data[j]),
                                                                                  str(Step_2_Data[j]))
                                continue
                            else:
                                ws.cell(row=i + 2, column=4).value = Rank_Data[j]
                                continue
                else:  # 1차유형에서 등급이 정해지지 않았고
                    for k in b:  # 2차유형에서 또한 여러항목과 중복되었다면
                        dic1[k] = Only_Char.find(k)  # 그 항목들 중 가장 마지막 항목을 기입하려함.
                    dic1_reverse = {v: k for k, v in dic1.items()}  # 딕셔너리와 find 메소도를 활용하여 인덱스를 구하고
                    value1 = ""  # 마지막 단어와 등급을 기입
                    for k, v in sorted(dic1_reverse.items()):
                        value1 = v
                        ws.cell(row=i + 2, column=2).value = value1 + value
                    for j in range(len(Step_1_Data)):
                        if Step_1_Data[j] == value and Step_2_Data[j] == value1:
                            if len(str(Rank_Data[j])) != 1:
                                print(Only_Char, Rank_Data[j])
                                ws.cell(row=i + 2, column=4).value = sep_specific(str(Only_Char), str(Step_1_Data[j]),
                                                                                  str(Step_2_Data[j]))
                            else:
                                ws.cell(row=i + 2, column=4).value = Rank_Data[j]

    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20
    wb.save(saveURL + '/result.xlsx')
    print("runaway end")
